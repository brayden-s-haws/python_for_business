'''
This is the importer used to read, analyze, and write data from CSV and Excel files
'''

import csv
from datetime import date
from openpyxl import load_workbook, Workbook

from importer import *

class FileImporter:

    def __init__(self, file_path, default_headers = None, default_data_types = None, row_data_type = ROW_TYPE_DICT, none_strings = DEFAULT_NONE_STRINGS, row_limit=None, row_filter=None):
        """A class used to read, clean, analyze, and write data from/to Excel and CSV
        :param str file_path: Name of the file to read
        :param dict default_headers: A key/value pair of header in raw file and the new name of header
        :param dict default_data_types: A key/value pair of header and the function or data type to convert each value to
        :param str row_data_type: A string indicating the data type that each row should be processed as
        :param list none_strings: A list of values that should be rewritten as a None value
        :param int row_limit: An integer that indicates how many row should be processed
        :param func row_filter: Allows you to filter data at time of import, limits imported data to records that fit filter parameters
        """
        self.default_headers = default_headers
        self.default_data_types = default_data_types
        self.row_data_type = row_data_type
        self.none_strings = none_strings

        self.file_reader = self.get_file_reader(file_path)
        self.headers = self.process_headers(list(next(self.file_reader)))
        print(f'Headers: {self.headers}')
        self.data = self.get_data(row_limit, row_filter)

    #### READ DATA ####
    def print_rows(self,num_rows):
        '''Allows you to print a specified number of rows after you have read in data
        :param int num_rows: The number of rows you wish to print
        :return: list
        '''
        for idx, row in enumerate(self.data):
            print(row)
            if idx == num_rows -1:
                break

    def get_data(self, row_limit, row_filter):
        """ Get all row data from the file import and return a processed list of records
        :return: list
        """
        data = []
        for idx, row in enumerate(self.file_reader):
            if row_limit and len(data) == row_limit:
                break
            processed_row = self.process_row(row)
            if idx == 0:
                headers_and_vals = zip(self.headers, processed_row) if self.row_data_type != ROW_TYPE_DICT else processed_row.items()
                for header, val in headers_and_vals:
                    print(f'{header} type is {type(val)}')
            if row_filter and not row_filter(processed_row):
                continue
            data.append(processed_row)
        # CSV file must be closed manually
        if self.file_type == FILE_TYPE_CSV:
            self.file.close()

        return data

    def process_row(self, row):
        """ Process the row into the data type specified by FileImporter
        :return: list|tuple|dict
        """
        processed_row = []
        for header, val in zip(self.headers, row):
            default_data_type = self.default_data_types.get(header) if self.default_data_types else None
            processed_row.append(self.process_value(val, default_data_type))
        if self.row_data_type == ROW_TYPE_LIST:
            return processed_row
        if self.row_data_type == ROW_TYPE_DICT:
            return {header: val for header, val in zip(self.headers, processed_row)}
        if self.row_data_type == ROW_TYPE_TUPLE:
            return tuple(processed_row)

    def process_value(self, val, default_data_type):
        """ Get a processed value. Handles None, null strings, and function conversions
        :param func|class defaultDataType: A function or class which converts the value into the appropriate format
        """
        if val is None:
            return val
        if isinstance(val, str):
            val = val.strip()
            if val.lower() in self.none_strings:
                return None
        if default_data_type:
            val = default_data_type(val)
        return val

    def get_file_reader(self,file_path):
        """ Get an iterator used to get each row in a CSV or XLSX file
        """
        if f'.{FILE_TYPE_CSV}' in file_path:
            self.file = open(file_path)
            self.file_type = FILE_TYPE_CSV
            return csv.reader(self.file)
        elif f'.{FILE_TYPE_XLS}' in file_path:
            self.file = load_workbook(file_path, data_only=True)
            worksheet = self.file.active
            self.file_type = FILE_TYPE_XLS
            return worksheet.iter_rows(values_only=True)
        else:
            raise(ValueError('Unsupported file type'))

    def process_headers(self, headers):
        """ Convert headers to formatted names
        """
        if not self.default_headers:
            return headers
        new_headers = []
        for idx, header in enumerate(headers):
            header_by_idx = self.default_headers.get(idx)
            header_by_value = self.default_headers.get(header)
            if header_by_idx:
                new_headers.append(header_by_idx)
            elif header_by_value:
                new_headers.append(header_by_value)
            else:
                new_headers.append(header)
        return new_headers

    #### WRITE DATA ####
    def write_csv_file(self, file_name, data=None, headers=None):
        """ Write data to a new CSV file
        :param str fileName: The name of the file to write to. The file extension should not be included.
        :param list data: If provided, will be used instead of the FileImporter's internal data property
        :param list headers: If provided, will be used instead of the FileImporter's internal header property
        """
        data_to_write = data or self.data
        headers_to_write = headers or self.headers

        with open(f'{DATA_FILE_OUTPUT_PATH}{file_name}_{date.today()}.csv', 'w') as csv_file:
            csv_writer = csv.writer(csv_file)
            csv_writer.writerow(headers_to_write)
            for row in data_to_write:
                csv_writer.writerow(self.format_row(row))
    def write_excel_file(self, file_name, sheets_config=None):
        """ Write data to a new XLSX file and return the workbook
        :param str fileName: The name of the file to write to. The file extension should not be included.
        :param list sheets_config: [{'title': , 'data': , 'headers':},...]
        """
        if not sheets_config:
            return

        workbook = Workbook()

        for idx, config in enumerate(sheets_config):
            # Create new worksheet
            if idx == 0:
                worksheet = workbook.active
            else:
                worksheet=workbook.create_sheet()
            worksheet.title = config.get('title', None)

            # Add data to sheet
            data_to_write = config.get('data', None) or self.data
            headers_to_write = config.get('headers', None) or self.headers
            worksheet.append(headers_to_write)
            for row in data_to_write:
                worksheet.append(self.format_row(row))

        # Save and return workbook
        workbook.save(f'{DATA_FILE_OUTPUT_PATH}{file_name}_{date.today()}.xlsx')
        return workbook


    def format_row(self,row):
        """ Process row to make sure it is an appropriate data format to write to a CSV or XSLX file.
        """
        if isinstance(row,(tuple,list)):
            return row
        if isinstance(row,dict):
            return list(row.values())
        raise(ValueError("Must use a row data type of tuple, list, or dict"))

    #### ANALYZE DATA ####
    def get_group_data(self, groupings,filter_fn=None, data=None, headers=None, is_flat=False):
        '''This allows you to group data based on arguments you pass in
        :param tuple groupings: the groupings for which you want to group data
        :param func filter_fn: function that allows you to filter data out of groupings
        :param dict data: the data you want to group
        :param dict headers: the headers for the groupings
        :param bool is_flat: allows you to specify if the data is flattened or not
        :return: dict
        '''
        data = data or self.data
        headers = headers or self.headers
        data_groups = {}
        for group in groupings:
            data_groups[self.get_group_column_key(group)] = {}

        for group in groupings:
            current_group = data_groups[self.get_group_column_key(group)]
            for record in data:
                if not filter_fn or filter_fn(record):
                    group_key = self.get_group_key(record, group, headers)
                    if group_key in current_group:
                        current_group[group_key].append(record)
                    else:
                        current_group[group_key] = [record]

        return data_groups if not is_flat else list(data_groups.values())

    def set_group_data(self,groupings):
        self.data_groups = self.get_group_data(groupings)

    def get_group_column_key(self,group):
        '''Returnd the column keys for groupings'''
        grouping_column_keys = []
        for column in group:
            if isinstance(column, str):
                grouping_column_keys.append(column)
            elif isinstance(column, tuple):
                grouping_column_keys.append(column[0])
            else:
                raise ValueError('Column must be string or tuple')
        return tuple(grouping_column_keys)

    def get_group_key(self, record, group, headers):
        '''Returns the key for a group'''
        if not isinstance(record, dict):
            record = {header: val for header, val in zip(headers, record)}

        group_key = []
        for column in group:
            if isinstance(column, tuple):
                column_key, grouping_func = column
                val = grouping_func(record[column_key])
            else:
                val = record[column]
            group_key.append(val)

        return tuple(group_key)